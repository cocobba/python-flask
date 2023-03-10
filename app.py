
##    return render_template("pay1.html",num1=num1,num2=num2,num3=num3,num4=num4[0],num5=num5)





import collections
from flask import Flask, render_template, request
from openpyxl import *



from flask import Flask, render_template, request
# from flask.ext.sqlalchemy import SQLAlchemy
import logging
from logging import Formatter, FileHandler
from forms import *
import os

#----------------------------------------------------------------------------#
# App Config.
#----------------------------------------------------------------------------#

app = Flask(__name__)
app.config.from_object('config')
#db = SQLAlchemy(app)

# Automatically tear down SQLAlchemy.
'''
@app.teardown_request
def shutdown_session(exception=None):
    db_session.remove()
'''

# Login required decorator.
'''
def login_required(test):
    @wraps(test)
    def wrap(*args, **kwargs):
        if 'logged_in' in session:
            return test(*args, **kwargs)
        else:
            flash('You need to login first.')
            return redirect(url_for('login'))
    return wrap
'''
#----------------------------------------------------------------------------#
# Controllers.
#----------------------------------------------------------------------------#




if not app.debug:
    file_handler = FileHandler('error.log')
    file_handler.setFormatter(
        Formatter('%(asctime)s %(levelname)s: %(message)s [in %(pathname)s:%(lineno)d]')
    )
    app.logger.setLevel(logging.INFO)
    file_handler.setLevel(logging.INFO)
    app.logger.addHandler(file_handler)
    app.logger.info('errors')

#----------------------------------------------------------------------------#

app = Flask(__name__)

@app.route('/',methods=["GET"])
def index():
    return render_template("index.html",a="      ")



@app.route('/home',methods=["GET"])
def home():
    
    global num6
    global num2
    global num3
    global num1
    global num5
    global numpay1
    global numpay2
    global numpay3
    
    
    num1=""
    num2=""
    num3=[]
    num4=""
    num5=""
    num=0
    edit=[]
    cnt = 0
    load = load_workbook("????????????1.xlsx", data_only=True)
    student = request.args.get("student")  
    load_sh=load['??????']
    for i in range(2,load_sh.max_row+1):
        numh=""
##        for k in range(7,11):
        numh=numh+str(load_sh['H'+str(i)].value)[7:]

        if student == numh:
            cnt = cnt + 1
            num1=load_sh['G'+str(i)].value  
            num2=load_sh['B'+str(i)].value  #?????? 
            numpay1=load_sh['I'+str(i)].value
            numpay2=load_sh['J'+str(i)].value
            numpay3=load_sh['K'+str(i)].value

            if num1 == None:
                num6='x'
            else:
                num6='o'
            for o in range(0,4):
                if o !=3:
                    num3.append(str(load_sh.cell(row=i,column=o+3).value))
                else:
                    num3.append(str(load_sh.cell(row=i,column=o+3).value))
                    
            for k in range(0,4):
                num=num+1
                if num3[k] == 'None':

                    num=num-1
                    edit.append(k)
            num5=str(num)
                
            numpay1=(str(numpay1).split(","))
            numpay2=(str(numpay2).split(","))
            numpay3=(str(numpay3).split(","))

            
            return render_template("lesson.html",num1=num6,num2=num2,num3=num3,num4=num1,num5=num5)
    if cnt == 0:
        return render_template("index.html",a="????????? ????????? ????????????")
    

@app.route('/name',methods=["GET"])
def name():
    global numpay1
    global numpay2
    global numpay3
    
    
    
    name=[]
    load = load_workbook("????????????1.xlsx", data_only=True)
    student = request.args.get("student")
    nm= request.args.get("name")
    load_sh=load['??????']

    

    for i in range(2,load_sh.max_row+1):
        numh=""    
        numh=numh+str(load_sh['H'+str(i)].value)[7:]
##        print(load_sh['B'+str(i)].value,'b')
        if load_sh['B'+str(i)].value==None:
           
##            print(student,'stu')
##            print(name,'na')
##            print(len(name),'len')
            if len(name)==0:
                return render_template("index.html",a="????????? ????????? ????????????")
            elif len(name)==1:
                home()
                return render_template("lesson.html",num1=num6,num2=num2,num3=num3,num4=num1,num5=num5)
            elif len(name)==2:
                name.append(None)
                return render_template("name.html",name1=name[0],name2=name[1],name3=name[2])
            else:
                return render_template("name.html",name1=name[0],name2=name[1],name3=name[2])
        elif student == numh:
            
            name.append(load_sh['B'+str(i)].value)
            #print(student,'stu')
##            
@app.route('/lesson')    
def lesson():
    global numpay1
    global numpay2
    global numpay3
    
    
##    print(request.args.get("kor")  )
    n1=""
    n2=""
    n3=[]
    n4=""
    n5=""
    n6=""
    n=0
    edit=[]
    
    load = load_workbook("????????????1.xlsx", data_only=True)
    student = request.args.get("student")
    # name= request.args.get("kor") ??????
    name= request.args.get("name")
    load_sh=load['??????']
##    print(name)


    for i in range(2,load_sh.max_row+1): 
        if load_sh['B'+str(i)].value == None:
            return render_template("lesson.html",num1=n6,num2=name,num3=n3,num4=n1,num5=n5)
            break
        if load_sh['B'+str(i)].value == name:
            n1=load_sh['G'+str(i)].value
            
            if n1 == None:
                n6='x'
            else:
                n6='o'
            for o in range(0,4):
                if o !=3:
                    n3.append(str(load_sh.cell(row=i,column=o+3).value))
                else:
                    n3.append(str(load_sh.cell(row=i,column=o+3).value))

                    
            for k in range(0,4):
                n=n+1
                if n3[k] == 'None':
                    n=n-1
                    edit.append(k)
            n5=str(n)
        if n1==load_sh['G'+str(i)].value and load_sh['B'+str(i)].value == name:
            print("d")
            numpay1=load_sh['I'+str(i)].value
            numpay2=load_sh['J'+str(i)].value
            numpay3=load_sh['K'+str(i)].value

            numpay1=(str(numpay1).split(","))
            numpay2=(str(numpay2).split(","))
            numpay3=(str(numpay3).split(","))
            return render_template("lesson.html",num1=n6,num2=name,num3=n3,num4=n1,num5=n5)

            
@app.route('/pay1')
def pay1():
    global numpay1
    global numpay2
    global numpay3
##    daylist=[]
##    c=find_name(num2)
##    a=(sh1_maxcolumn(c))//5
##    for i in range(1,a+1):
##        daylist.append(load_sh1(row=c,column=(a*i)+1))
##        daylist.append(load_sh1(row=c,column=(a*i)+2))
##        daylist.append(load_sh1(row=c,column=(a*i)+3))
##        daylist.append(load_sh1(row=c,column=(a*i)+4))
#1  
    if numpay1[0] != "None":
        
        n1 = numpay1[0][1:len(numpay1[0])]
        n2 = numpay1[1][0:len(numpay1[1])]
        n3 = numpay1[2][0:len(numpay1[2])]
        n4 = numpay1[3][0:len(numpay1[3])]
        n5 = numpay1[4][0:len(numpay1[4])-1]
    else:
        n1 ='None'
        n2 = 'None'
        n3 = 'None'
        n4 = 'None'
        n5 = 'None'
    
#2
    if numpay2[0] != "None":
        
        n6 = numpay2[0][1:len(numpay2[0])]
        n7 = numpay2[1][0:len(numpay2[1])]
        n8 = numpay2[2][0:len(numpay2[2])]
        n9 = numpay2[3][0:len(numpay2[3])]
        n10 = numpay2[4][0:len(numpay2[4])-1]
    else:
        n6 = 'None'
        n7 = 'None'
        n8 = 'None'
        n9 = 'None'
        n10 = 'None'
#3
    if numpay3[0] != "None":

        n11 = numpay3[0][1:len(numpay3[0])]
        n12= numpay3[1][0:len(numpay3[1])]
        
        n13= numpay3[2][0:len(numpay3[2])]
        n14= numpay3[3][0:len(numpay3[3])]
        n15 = numpay3[4][0:len(numpay3[4])-1]
    else:
##        print(numpay3,'32')
        n11 = 'None'
        n12= 'None'
        n13= 'None'
        n14= 'None'
        n15 = 'None'
        
    return render_template("pay1.html",num1=n1,num2=n2,num3=n3,num4=n4,num5=n5,num6=n6,num7=n7,num8=n8,num9=n9,num10=n10,num11=n11,num12=n12,num13=n13,num14=n14,num15=n15)
# Launch.
#----------------------------------------------------------------------------#

# Default port:
if __name__ == '__main__':
    app.run()

